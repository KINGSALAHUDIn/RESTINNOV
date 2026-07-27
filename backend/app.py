from flask import Flask, request, jsonify, send_file,send_from_directory
from flask_cors import CORS
from groq import Groq
import os
import io
import tempfile
import json
import sqlite3
from dotenv import load_dotenv
import threading
import webbrowser
# ── Excel generation ──────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

load_dotenv()


app = Flask(
    __name__,
    static_folder="dist",
    static_url_path=""
)
@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def catch_all(path):
    return send_from_directory(app.static_folder, "index.html")
CORS(app, resources={r"/*": {"origins": "*"}})

api_key = os.environ.get("GROQ_API_KEY")
if not api_key:
    print("ERROR: GROQ_API_KEY not found in .env")
else:
    print(f"Groq API key loaded: ...{api_key[-4:]}")

client = Groq(api_key=api_key)

DB_PATH = "inspections.db"

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS appartements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date_inspection TEXT,
            proprietaire TEXT,
            controleur TEXT,
            code_appartement TEXT,
            adresse TEXT,
            transcript TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS entree (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            appartement_id INTEGER,
            porte_entree_conforme INTEGER, porte_entree_commentaire TEXT,
            sol_conforme INTEGER, sol_commentaire TEXT,
            murs_conforme INTEGER, murs_commentaire TEXT,
            lumiere_conforme INTEGER, lumiere_commentaire TEXT,
            interphone_digicode_conforme INTEGER, interphone_digicode_commentaire TEXT,
            odeur_conforme INTEGER, odeur_commentaire TEXT,
            FOREIGN KEY(appartement_id) REFERENCES appartements(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS salon (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            appartement_id INTEGER,
            sol_murs_conforme INTEGER, sol_murs_commentaire TEXT,
            poussiere_meubles_conforme INTEGER, poussiere_meubles_commentaire TEXT,
            canape_conforme INTEGER, canape_commentaire TEXT,
            coussins_conforme INTEGER, coussins_commentaire TEXT,
            table_basse_conforme INTEGER, table_basse_commentaire TEXT,
            rideaux_conforme INTEGER, rideaux_commentaire TEXT,
            volets_conforme INTEGER, volets_commentaire TEXT,
            fenetres_conforme INTEGER, fenetres_commentaire TEXT,
            eclairage_conforme INTEGER, eclairage_commentaire TEXT,
            television_conforme INTEGER, television_commentaire TEXT,
            decoration_conforme INTEGER, decoration_commentaire TEXT,
            odeur_conforme INTEGER, odeur_commentaire TEXT,
            FOREIGN KEY(appartement_id) REFERENCES appartements(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS cuisine (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            appartement_id INTEGER,
            sol_murs_conforme INTEGER, sol_murs_commentaire TEXT,
            plan_travail_conforme INTEGER, plan_travail_commentaire TEXT,
            evier_conforme INTEGER, evier_commentaire TEXT,
            vaisselle_conforme INTEGER, vaisselle_commentaire TEXT,
            ustensiles_conforme INTEGER, ustensiles_commentaire TEXT,
            plaques_cuisson_conforme INTEGER, plaques_cuisson_commentaire TEXT,
            hotte_conforme INTEGER, hotte_commentaire TEXT,
            micro_ondes_four_conforme INTEGER, micro_ondes_four_commentaire TEXT,
            refrigerateur_conforme INTEGER, refrigerateur_commentaire TEXT,
            machine_cafe_bouilloire_conforme INTEGER, machine_cafe_bouilloire_commentaire TEXT,
            placards_conforme INTEGER, placards_commentaire TEXT,
            poubelle_conforme INTEGER, poubelle_commentaire TEXT,
            produits_menagers_conforme INTEGER, produits_menagers_commentaire TEXT,
            FOREIGN KEY(appartement_id) REFERENCES appartements(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS salle_de_bain (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            appartement_id INTEGER,
            sol_murs_conforme INTEGER, sol_murs_commentaire TEXT,
            lavabo_conforme INTEGER, lavabo_commentaire TEXT,
            miroir_conforme INTEGER, miroir_commentaire TEXT,
            douche_baignoire_conforme INTEGER, douche_baignoire_commentaire TEXT,
            wc_conforme INTEGER, wc_commentaire TEXT,
            papier_toilette_conforme INTEGER, papier_toilette_commentaire TEXT,
            poubelle_conforme INTEGER, poubelle_commentaire TEXT,
            lumieres_conforme INTEGER, lumieres_commentaire TEXT,
            odeur_conforme INTEGER, odeur_commentaire TEXT,
            FOREIGN KEY(appartement_id) REFERENCES appartements(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS chambres (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            appartement_id INTEGER,
            sol_murs_conforme INTEGER, sol_murs_commentaire TEXT,
            poussiere_conforme INTEGER, poussiere_commentaire TEXT,
            literie_conforme INTEGER, literie_commentaire TEXT,
            lit_conforme INTEGER, lit_commentaire TEXT,
            matelas_conforme INTEGER, matelas_commentaire TEXT,
            placard_dressing_conforme INTEGER, placard_dressing_commentaire TEXT,
            lumieres_conforme INTEGER, lumieres_commentaire TEXT,
            television_conforme INTEGER, television_commentaire TEXT,
            rideaux_volets_conforme INTEGER, rideaux_volets_commentaire TEXT,
            odeur_conforme INTEGER, odeur_commentaire TEXT,
            FOREIGN KEY(appartement_id) REFERENCES appartements(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS balcon_terrasse (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            appartement_id INTEGER,
            sol_murs_conforme INTEGER, sol_murs_commentaire TEXT,
            mobilier_exterieur_conforme INTEGER, mobilier_exterieur_commentaire TEXT,
            proprete_conforme INTEGER, proprete_commentaire TEXT,
            garde_corps_conforme INTEGER, garde_corps_commentaire TEXT,
            FOREIGN KEY(appartement_id) REFERENCES appartements(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS equipements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            appartement_id INTEGER,
            interrupteurs_conforme INTEGER, interrupteurs_commentaire TEXT,
            prises_conforme INTEGER, prises_commentaire TEXT,
            wifi_conforme INTEGER, wifi_commentaire TEXT,
            eau_chaude_conforme INTEGER, eau_chaude_commentaire TEXT,
            chauffage_conforme INTEGER, chauffage_commentaire TEXT,
            climatisation_conforme INTEGER, climatisation_commentaire TEXT,
            machine_laver_conforme INTEGER, machine_laver_commentaire TEXT,
            portes_serrures_conforme INTEGER, portes_serrures_commentaire TEXT,
            cles_conforme INTEGER, cles_commentaire TEXT,
            FOREIGN KEY(appartement_id) REFERENCES appartements(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS controle_sensoriel (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            appartement_id INTEGER,
            odeur_generale_conforme INTEGER, odeur_generale_commentaire TEXT,
            temperature_conforme INTEGER, temperature_commentaire TEXT,
            bruit_conforme INTEGER, bruit_commentaire TEXT,
            lumiere_ambiante_conforme INTEGER, lumiere_ambiante_commentaire TEXT,
            FOREIGN KEY(appartement_id) REFERENCES appartements(id)
        )
    """)
    conn.commit()
    conn.close()
    print("SQLite database initialized")

init_db()

SYSTEM_PROMPT = """Tu es un assistant qui extrait des informations d'un rapport d'inspection d'appartement.
A partir du texte donne, extrais toutes les informations et retourne UNIQUEMENT un JSON valide, sans texte supplementaire, sans balises markdown.

Structure JSON exacte:
{
  "date": "",
  "proprietaire": "",
  "controleur": "",
  "code_appartement": "",
  "adresse": "",
  "sections": {
    "entree": {"porte_entree": {"conforme": null, "commentaire": ""}, "sol": {"conforme": null, "commentaire": ""}, "murs": {"conforme": null, "commentaire": ""}, "lumiere": {"conforme": null, "commentaire": ""}, "interphone_digicode": {"conforme": null, "commentaire": ""}, "odeur": {"conforme": null, "commentaire": ""}},
    "salon": {"sol_murs": {"conforme": null, "commentaire": ""}, "poussiere_meubles": {"conforme": null, "commentaire": ""}, "canape": {"conforme": null, "commentaire": ""}, "coussins": {"conforme": null, "commentaire": ""}, "table_basse": {"conforme": null, "commentaire": ""}, "rideaux": {"conforme": null, "commentaire": ""}, "volets": {"conforme": null, "commentaire": ""}, "fenetres": {"conforme": null, "commentaire": ""}, "eclairage": {"conforme": null, "commentaire": ""}, "television": {"conforme": null, "commentaire": ""}, "decoration": {"conforme": null, "commentaire": ""}, "odeur": {"conforme": null, "commentaire": ""}},
    "cuisine": {"sol_murs": {"conforme": null, "commentaire": ""}, "plan_travail": {"conforme": null, "commentaire": ""}, "evier": {"conforme": null, "commentaire": ""}, "vaisselle": {"conforme": null, "commentaire": ""}, "ustensiles": {"conforme": null, "commentaire": ""}, "plaques_cuisson": {"conforme": null, "commentaire": ""}, "hotte": {"conforme": null, "commentaire": ""}, "micro_ondes_four": {"conforme": null, "commentaire": ""}, "refrigerateur": {"conforme": null, "commentaire": ""}, "machine_cafe_bouilloire": {"conforme": null, "commentaire": ""}, "placards": {"conforme": null, "commentaire": ""}, "poubelle": {"conforme": null, "commentaire": ""}, "produits_menagers": {"conforme": null, "commentaire": ""}},
    "salle_de_bain": {"sol_murs": {"conforme": null, "commentaire": ""}, "lavabo": {"conforme": null, "commentaire": ""}, "miroir": {"conforme": null, "commentaire": ""}, "douche_baignoire": {"conforme": null, "commentaire": ""}, "wc": {"conforme": null, "commentaire": ""}, "papier_toilette": {"conforme": null, "commentaire": ""}, "poubelle": {"conforme": null, "commentaire": ""}, "lumieres": {"conforme": null, "commentaire": ""}, "odeur": {"conforme": null, "commentaire": ""}},
    "chambres": {"sol_murs": {"conforme": null, "commentaire": ""}, "poussiere": {"conforme": null, "commentaire": ""}, "literie": {"conforme": null, "commentaire": ""}, "lit": {"conforme": null, "commentaire": ""}, "matelas": {"conforme": null, "commentaire": ""}, "placard_dressing": {"conforme": null, "commentaire": ""}, "lumieres": {"conforme": null, "commentaire": ""}, "television": {"conforme": null, "commentaire": ""}, "rideaux_volets": {"conforme": null, "commentaire": ""}, "odeur": {"conforme": null, "commentaire": ""}},
    "balcon_terrasse": {"sol_murs": {"conforme": null, "commentaire": ""}, "mobilier_exterieur": {"conforme": null, "commentaire": ""}, "proprete": {"conforme": null, "commentaire": ""}, "garde_corps": {"conforme": null, "commentaire": ""}},
    "equipements": {"interrupteurs": {"conforme": null, "commentaire": ""}, "prises": {"conforme": null, "commentaire": ""}, "wifi": {"conforme": null, "commentaire": ""}, "eau_chaude": {"conforme": null, "commentaire": ""}, "chauffage": {"conforme": null, "commentaire": ""}, "climatisation": {"conforme": null, "commentaire": ""}, "machine_laver": {"conforme": null, "commentaire": ""}, "portes_serrures": {"conforme": null, "commentaire": ""}, "cles": {"conforme": null, "commentaire": ""}},
    "controle_sensoriel": {"odeur_generale": {"conforme": null, "commentaire": ""}, "temperature": {"conforme": null, "commentaire": ""}, "bruit": {"conforme": null, "commentaire": ""}, "lumiere_ambiante": {"conforme": null, "commentaire": ""}}
  }
}

conforme: true si conforme, false si non conforme, null si non mentionne
Retourne UNIQUEMENT le JSON, rien d autre"""


# ── Excel helper ──────────────────────────────────────────────────────────────

SECTION_LABELS = {
    "entree":             "Entrée",
    "salon":              "Salon",
    "cuisine":            "Cuisine",
    "salle_de_bain":      "Salle de Bain",
    "chambres":           "Chambres",
    "balcon_terrasse":    "Balcon / Terrasse",
    "equipements":        "Équipements",
    "controle_sensoriel": "Contrôle Sensoriel",
}

ITEM_LABELS = {
    "porte_entree": "Porte d'entrée", "sol": "Sol", "murs": "Murs",
    "lumiere": "Lumière", "interphone_digicode": "Interphone / Digicode",
    "odeur": "Odeur", "sol_murs": "Sol & Murs",
    "poussiere_meubles": "Poussière / Meubles", "canape": "Canapé",
    "coussins": "Coussins", "table_basse": "Table basse", "rideaux": "Rideaux",
    "volets": "Volets", "fenetres": "Fenêtres", "eclairage": "Éclairage",
    "television": "Télévision", "decoration": "Décoration",
    "plan_travail": "Plan de travail", "evier": "Évier", "vaisselle": "Vaisselle",
    "ustensiles": "Ustensiles", "plaques_cuisson": "Plaques de cuisson",
    "hotte": "Hotte", "micro_ondes_four": "Micro-ondes / Four",
    "refrigerateur": "Réfrigérateur",
    "machine_cafe_bouilloire": "Machine à café / Bouilloire",
    "placards": "Placards", "poubelle": "Poubelle",
    "produits_menagers": "Produits ménagers", "lavabo": "Lavabo",
    "miroir": "Miroir", "douche_baignoire": "Douche / Baignoire",
    "wc": "WC", "papier_toilette": "Papier toilette", "lumieres": "Lumières",
    "poussiere": "Poussière", "literie": "Literie", "lit": "Lit",
    "matelas": "Matelas", "placard_dressing": "Placard / Dressing",
    "rideaux_volets": "Rideaux / Volets",
    "mobilier_exterieur": "Mobilier extérieur", "proprete": "Propreté",
    "garde_corps": "Garde-corps", "interrupteurs": "Interrupteurs",
    "prises": "Prises", "wifi": "WiFi", "eau_chaude": "Eau chaude",
    "chauffage": "Chauffage", "climatisation": "Climatisation",
    "machine_laver": "Machine à laver", "portes_serrures": "Portes / Serrures",
    "cles": "Clés", "odeur_generale": "Odeur générale",
    "temperature": "Température", "bruit": "Bruit",
    "lumiere_ambiante": "Lumière ambiante",
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _apply(cell, font=None, fill=None, align=None):
    if font:  cell.font      = font
    if fill:  cell.fill      = fill
    if align: cell.alignment = align
    cell.border = _border()

def generate_inspection_excel(inspection: dict, meta: dict = None) -> bytes:
    BLUE_DARK  = "1F3864"
    BLUE_MID   = "2E75B6"
    BLUE_LIGHT = "D6E4F0"
    WHITE      = "FFFFFF"
    GREEN      = "C6EFCE"
    RED        = "FFC7CE"
    GREY       = "F2F2F2"

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Inspection"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 55
    ws.freeze_panes = "A5"

    row = 1

    # Title
    ws.merge_cells(f"A{row}:C{row}")
    c = ws[f"A{row}"]
    c.value = "RAPPORT D'INSPECTION D'APPARTEMENT"
    _apply(c, font=Font(name="Arial", bold=True, size=14, color=WHITE),
           fill=_fill(BLUE_DARK),
           align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[row].height = 28
    row += 1

    # Meta fields
    meta_fields = [
        ("Date d'inspection", inspection.get("date", "") or (meta or {}).get("date_inspection", "")),
        ("Propriétaire",      inspection.get("proprietaire", "")),
        ("Contrôleur",        inspection.get("controleur", "")),
        ("Code appartement",  inspection.get("code_appartement", "")),
        ("Adresse",           inspection.get("adresse", "")),
    ]
    if meta and meta.get("id"):
        meta_fields.append(("ID base de données", str(meta["id"])))

    for label, value in meta_fields:
        ws.merge_cells(f"A{row}:C{row}")
        lc = ws[f"A{row}"]
        lc.value = f"  {label} :   {value}"
        _apply(lc, font=Font(name="Arial", bold=True, size=9),
               fill=_fill(BLUE_LIGHT),
               align=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1

    # Column headers
    for col_idx, label in enumerate(["Élément", "Conforme ?", "Commentaire"], 1):
        hc = ws.cell(row=row, column=col_idx, value=label)
        _apply(hc, font=Font(name="Arial", bold=True, size=10, color=WHITE),
               fill=_fill(BLUE_MID),
               align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[row].height = 20
    row += 1

    sections = inspection.get("sections", {})
    alt = False
    total = conf_count = nonconf_count = 0

    for section_key, section_label in SECTION_LABELS.items():
        items = sections.get(section_key, {})
        if not items:
            continue

        # Section header
        ws.merge_cells(f"A{row}:C{row}")
        sc = ws[f"A{row}"]
        sc.value = f"  {section_label.upper()}"
        _apply(sc, font=Font(name="Arial", bold=True, size=10, color=WHITE),
               fill=_fill(BLUE_MID),
               align=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[row].height = 18
        row += 1

        for item_key, item_data in items.items():
            conforme    = item_data.get("conforme")    if isinstance(item_data, dict) else None
            commentaire = item_data.get("commentaire", "") if isinstance(item_data, dict) else ""

            bg = BLUE_LIGHT if alt else WHITE
            alt = not alt

            if conforme is True:
                conf_text = "✔  Conforme"
                conf_fill = _fill(GREEN)
                conf_count += 1; total += 1
            elif conforme is False:
                conf_text = "✘  Non conforme"
                conf_fill = _fill(RED)
                nonconf_count += 1; total += 1
            else:
                conf_text = "—"
                conf_fill = _fill(GREY)

            label_text = ITEM_LABELS.get(item_key, item_key.replace("_", " ").title())

            ca = ws.cell(row=row, column=1, value=f"  {label_text}")
            cb = ws.cell(row=row, column=2, value=conf_text)
            cc = ws.cell(row=row, column=3, value=commentaire)

            _apply(ca, font=Font(name="Arial", size=9), fill=_fill(bg),
                   align=Alignment(horizontal="left", vertical="center"))
            _apply(cb, font=Font(name="Arial", bold=True, size=9), fill=conf_fill,
                   align=Alignment(horizontal="center", vertical="center"))
            _apply(cc, font=Font(name="Arial", size=9), fill=_fill(bg),
                   align=Alignment(horizontal="left", vertical="center", wrap_text=True))
            ws.row_dimensions[row].height = 15
            row += 1

        row += 1  # spacer between sections

    # Summary
    ws.merge_cells(f"A{row}:C{row}")
    sm = ws[f"A{row}"]
    sm.value = (
        f"  RÉSUMÉ — Éléments évalués : {total}   |   "
        f"Conformes : {conf_count}   |   Non conformes : {nonconf_count}"
    )
    _apply(sm, font=Font(name="Arial", bold=True, size=10, color=WHITE),
           fill=_fill(BLUE_DARK),
           align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[row].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── DB helpers ────────────────────────────────────────────────────────────────

def _b(val):
    """Convert Python bool / None → SQLite int / None."""
    if val is True:  return 1
    if val is False: return 0
    return None

def save_to_db(inspection_data, transcript):
    conn = get_db()
    c = conn.cursor()

    c.execute("""
        INSERT INTO appartements (date_inspection, proprietaire, controleur, code_appartement, adresse, transcript)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (
        inspection_data.get("date", ""),
        inspection_data.get("proprietaire", ""),
        inspection_data.get("controleur", ""),
        inspection_data.get("code_appartement", ""),
        inspection_data.get("adresse", ""),
        transcript
    ))
    apt_id = c.lastrowid
    s = inspection_data.get("sections", {})

    def g(section, key):
        return s.get(section, {}).get(key, {})

    if "entree" in s:
        e = s["entree"]
        c.execute("""INSERT INTO entree
            (appartement_id,
             porte_entree_conforme, porte_entree_commentaire,
             sol_conforme, sol_commentaire,
             murs_conforme, murs_commentaire,
             lumiere_conforme, lumiere_commentaire,
             interphone_digicode_conforme, interphone_digicode_commentaire,
             odeur_conforme, odeur_commentaire)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (apt_id,
             _b(g("entree","porte_entree").get("conforme")), g("entree","porte_entree").get("commentaire",""),
             _b(g("entree","sol").get("conforme")),          g("entree","sol").get("commentaire",""),
             _b(g("entree","murs").get("conforme")),         g("entree","murs").get("commentaire",""),
             _b(g("entree","lumiere").get("conforme")),      g("entree","lumiere").get("commentaire",""),
             _b(g("entree","interphone_digicode").get("conforme")), g("entree","interphone_digicode").get("commentaire",""),
             _b(g("entree","odeur").get("conforme")),        g("entree","odeur").get("commentaire","")))

    if "salon" in s:
        keys = ["sol_murs","poussiere_meubles","canape","coussins","table_basse",
                "rideaux","volets","fenetres","eclairage","television","decoration","odeur"]
        cols = ",".join(f"{k}_conforme,{k}_commentaire" for k in keys)
        vals = []
        for k in keys:
            vals += [_b(g("salon",k).get("conforme")), g("salon",k).get("commentaire","")]
        c.execute(f"INSERT INTO salon (appartement_id,{cols}) VALUES (?{',?' * len(vals)})",
                  [apt_id] + vals)

    if "cuisine" in s:
        keys = ["sol_murs","plan_travail","evier","vaisselle","ustensiles","plaques_cuisson",
                "hotte","micro_ondes_four","refrigerateur","machine_cafe_bouilloire",
                "placards","poubelle","produits_menagers"]
        cols = ",".join(f"{k}_conforme,{k}_commentaire" for k in keys)
        vals = []
        for k in keys:
            vals += [_b(g("cuisine",k).get("conforme")), g("cuisine",k).get("commentaire","")]
        c.execute(f"INSERT INTO cuisine (appartement_id,{cols}) VALUES (?{',?' * len(vals)})",
                  [apt_id] + vals)

    if "salle_de_bain" in s:
        keys = ["sol_murs","lavabo","miroir","douche_baignoire","wc",
                "papier_toilette","poubelle","lumieres","odeur"]
        cols = ",".join(f"{k}_conforme,{k}_commentaire" for k in keys)
        vals = []
        for k in keys:
            vals += [_b(g("salle_de_bain",k).get("conforme")), g("salle_de_bain",k).get("commentaire","")]
        c.execute(f"INSERT INTO salle_de_bain (appartement_id,{cols}) VALUES (?{',?' * len(vals)})",
                  [apt_id] + vals)

    if "chambres" in s:
        keys = ["sol_murs","poussiere","literie","lit","matelas",
                "placard_dressing","lumieres","television","rideaux_volets","odeur"]
        cols = ",".join(f"{k}_conforme,{k}_commentaire" for k in keys)
        vals = []
        for k in keys:
            vals += [_b(g("chambres",k).get("conforme")), g("chambres",k).get("commentaire","")]
        c.execute(f"INSERT INTO chambres (appartement_id,{cols}) VALUES (?{',?' * len(vals)})",
                  [apt_id] + vals)

    if "balcon_terrasse" in s:
        keys = ["sol_murs","mobilier_exterieur","proprete","garde_corps"]
        cols = ",".join(f"{k}_conforme,{k}_commentaire" for k in keys)
        vals = []
        for k in keys:
            vals += [_b(g("balcon_terrasse",k).get("conforme")), g("balcon_terrasse",k).get("commentaire","")]
        c.execute(f"INSERT INTO balcon_terrasse (appartement_id,{cols}) VALUES (?{',?' * len(vals)})",
                  [apt_id] + vals)

    if "equipements" in s:
        keys = ["interrupteurs","prises","wifi","eau_chaude","chauffage",
                "climatisation","machine_laver","portes_serrures","cles"]
        cols = ",".join(f"{k}_conforme,{k}_commentaire" for k in keys)
        vals = []
        for k in keys:
            vals += [_b(g("equipements",k).get("conforme")), g("equipements",k).get("commentaire","")]
        c.execute(f"INSERT INTO equipements (appartement_id,{cols}) VALUES (?{',?' * len(vals)})",
                  [apt_id] + vals)

    if "controle_sensoriel" in s:
        keys = ["odeur_generale","temperature","bruit","lumiere_ambiante"]
        cols = ",".join(f"{k}_conforme,{k}_commentaire" for k in keys)
        vals = []
        for k in keys:
            vals += [_b(g("controle_sensoriel",k).get("conforme")), g("controle_sensoriel",k).get("commentaire","")]
        c.execute(f"INSERT INTO controle_sensoriel (appartement_id,{cols}) VALUES (?{',?' * len(vals)})",
                  [apt_id] + vals)

    conn.commit()
    conn.close()
    return apt_id


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/transcribe", methods=["POST"])
def transcribe():
    print("Received inspection audio")

    if "audio" not in request.files:
        return jsonify({"error": "No audio"}), 400

    audio_file = request.files["audio"]
    with tempfile.NamedTemporaryFile(suffix=".webm", delete=False) as tmp:
        audio_file.save(tmp.name)
        tmp_path = tmp.name

    try:
        print("Transcribing...")
        with open(tmp_path, "rb") as f:
            whisper_response = client.audio.transcriptions.create(
                model="whisper-large-v3",
                file=f,
                language="fr"
            )
        transcript = whisper_response.text
        print(f"Transcript: {transcript[:100]}")
    except Exception as e:
        print(f"Whisper error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        os.remove(tmp_path)

    try:
        print("Extracting JSON...")
        llm_response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": transcript}
            ],
            temperature=0.1,
        )
        raw = llm_response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        raw = raw.strip()
        inspection_data = json.loads(raw)
        print("JSON OK")
    except Exception as e:
        print(f"LLM error: {e}")
        return jsonify({"error": str(e)}), 500

    try:
        print("Saving to DB...")
        appartement_id = save_to_db(inspection_data, transcript)
        print(f"Saved ID: {appartement_id}")
    except Exception as e:
        print(f"DB error: {e}")
        return jsonify({"transcript": transcript, "inspection": inspection_data, "saved": False, "db_error": str(e)})

    return jsonify({"transcript": transcript, "inspection": inspection_data, "saved": True, "id": appartement_id})


@app.route("/export/<int:appartement_id>", methods=["GET"])
def export_excel(appartement_id):
    """
    Generate and stream an Excel inspection report for the given appartement ID.
    The inspection JSON is rebuilt from the data passed as a query-param (base64 JSON)
    OR fetched from the DB transcript column.
    
    Simplest approach: the frontend sends the inspection JSON it already has
    via a POST body — no need to re-query all joined tables.
    """
    return jsonify({"error": "Use POST /export"}), 405


@app.route("/export", methods=["POST"])
def export_excel_post():
    """
    Accepts JSON body: { "inspection": {...}, "id": 42 }
    Returns an .xlsx file download.
    """
    body = request.get_json(force=True, silent=True) or {}
    inspection = body.get("inspection")
    apt_id     = body.get("id")

    if not inspection:
        return jsonify({"error": "Missing inspection data"}), 400

    meta = {"id": apt_id} if apt_id else {}

    try:
        xlsx_bytes = generate_inspection_excel(inspection, meta)
    except Exception as e:
        print(f"Excel generation error: {e}")
        return jsonify({"error": str(e)}), 500

    code = inspection.get("code_appartement") or str(apt_id or "inspection")
    filename = f"inspection_{code}.xlsx"

    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

if __name__ == "__main__":

    
   app.run(host="0.0.0.0", port=5000)
